import sys
import getpass
import platform
import time
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter

def checkAndLoad(path):
    try:
        wb = load_workbook(path)
        ws = wb.active #active means last opened sheet
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    return [wb,ws]

def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving file")
    
    print("File is successfully saved in specified file path.")

def writeLog(ws):
    name = getpass.getuser().lower()
    timestamp = time.strftime("%b %d %Y %I:%M %p")
    max_row,row_loc  = ws.max_row, 2
    print(max_row)
    if max_row == 1:
        ws['A2']= name
        ws['B2']= 1
        ws['C2']= timestamp
        return
    

    for row in ws.iter_rows(min_row=2, max_row=50, min_col=1, max_col=3, values_only=True):
        if row[0]:
            if name == row[0]:
                ws.cell(row=row_loc, column=2).value = row[1] + 1 if isinstance(row[1],int) else 1
                ws.cell(row=row_loc, column=3).value = timestamp
                print(row_loc)
                return
        row_loc += 1

    maxim1 = str(max_row+1)
    ws['A'+maxim1] = name
    ws['B'+maxim1] = 1
    ws['C'+maxim1] = timestamp
    


def main():
    system = platform.system() ##Windows or MAC

    path = r'/Users/sebastienstvil/Desktop/log.xlsx'\
         if system == 'Darwin' else r'C:\Users\stvils\OneDrive - Wellington Management\Documents\GA Lab\GA Lab\Python-DEV\Userlog.xlsx'

    
    wbws = checkAndLoad(path)
    wb = wbws[0]
    ws = wbws[1]

    try:
        writeLog(ws)
    except Exception:
        print("Error in writting file")


    checkAndSave(wb,path)
    

    
    

if __name__ == "__main__":
    main()