from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font

##Transforms array of Strings into consistent characters and case
def arrangeStrings(array):
    for i in range(len(array)-1):
        if array[i]:
            array[i] = array[i].strip().replace(" ", "_").upper()
    return array 

#Returns the Total amount for the Complex ALpha pool in BOS
def get_complex_alpha_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "COMPLEX_ALPHA" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_core_equity_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "CORE_EQUITY" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_fixed_income_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "FIXED_INCOME" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_hk(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "HK" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_sf(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "SF" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_sing(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "SING" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_tok(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "TOK" in array[0] :
                amount = amount + array[1]
    return amount

def get_special_equity(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "SPECIAL_EQUITY" and ("BOS" in array[0] or "RAD" in array[0]):
                #Since we might surpas the rows which contain values, we're ensuring that we're only counting numbers
                if array[1]:
                    amount = amount + array[1]
    return amount

def get_global_domestic(dic):
    return get_global_bos(dic) + get_global_sf(dic) - GLOBAL_OPPORTUNISTIC

def calc_isvalid():
    CHECK_TOTAL = get_core_equity_bos(dic) + get_complex_alpha_bos(dic) + get_fixed_income_bos(dic) + get_special_equity(dic) \
        + get_global_domestic(dic) + GLOBAL_OPPORTUNISTIC + get_global_tok(dic) + get_global_sing(dic) + get_global_hk(dic)
    return CHECK_TOTAL == GRAND_TOTAL

def changeWidth(ws, beg_col, size):
    column = beg_col
    while column < 601:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = size
        column += 1

    #Change number to comma-style
    for row in range(1, 100):
        ws["E{}".format(row)].number_format = '#,##0.00_-'
        ws["F{}".format(row)].number_format = '#,##0.00_-'
        ws["G{}".format(row)].number_format = '#,##0.00_-'
        ws["N{}".format(row)].number_format = '#,##0.00_-'
   # ecell.number_format = 

def addtofile(COMPLEX_ALPHA, CORE_EQUITY, FIXED_INCOME, SPECIAL_EQUITY, 
            GLOBAL_DOMESTIC, GLOBAL_OPPORTUNISTIC, GLOBAL_TOKYO,\
            GLOBAL_SINGAPORE, GLOBAL_HONGKONG, ws):
    
    result = {'COMPLEX_ALPHA': COMPLEX_ALPHA, 'CORE_EQUITY': CORE_EQUITY, 'FIXED_INCOME': FIXED_INCOME, 'SPECIAL_EQUITY': SPECIAL_EQUITY,
    'GLOBAL_INTL_EQUITY_DOMESTIC': GLOBAL_DOMESTIC, 'GLOBAL_INTL_EQUITY_OPPORTUNISTIC': GLOBAL_OPPORTUNISTIC, 'GLOBAL INTL_EQUITY_TOKYO': GLOBAL_TOKYO,\
        'GLOBAL_INTL_EQUITY_SINGAPORE': GLOBAL_SINGAPORE, 'GLOBAL_INTL_EQUITY_HONGKONG':GLOBAL_HONGKONG}
    
    total = 0
    starting_row, starting_col = 8, 12
    for k, v in result.items():
        total += v
        if 'domestic' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Domestic'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'opportunistic' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Opportunistic'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'tokyo' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Tokyo'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'singapore' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Singapore'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'hong' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Hong Kong'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        else:
            ws.cell(row=starting_row, column=starting_col).value = k
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
    
    #At this stage, this is pointing to the end row
    starting_row += 2
    ws.cell(row=starting_row, column=starting_col).value = 'Total'
    ws['N19'] = '=SUM(N8:N16)'
     

def saveworkbook():
    path = '/Users/sebastienstvil/Documents/Python/Python-DEV/GA Lab/Bucket Report/bucketRes.xlsx'
    wb.save(path)

#Load existing spreadsheet

source = '/Users/sebastienstvil/Documents/Python/Python-DEV/GA Lab/Bucket Report/Bucket Report(raw).xlsx'
wb = load_workbook(source)

#Create active worksheet
ws = wb.active

#We use a dictionnary to chain the key(pool Bucket) to values
dic = {}
GLOBAL_OPPORTUNISTIC, GRAND_TOTAL = 0, 0
GRAND_TOTAL_row , i = 0, 0
loc = []
pool = "" 
for row in ws.iter_rows(min_row=8, max_row=300, min_col=1, max_col=5, values_only=True):
    print(row)
    if row[0]:
        if 'grand total' in row[0].lower():
            GRAND_TOTAL += row[4]
    if 'Total'.upper() not in str(row[0]).upper(): # as long as the first row does not contain the word "Total" 
        #Upper is used for case Insensitive Comparison)
        lst = []

        #Get Opportunistic before "Acoount Names" are unavailable when tuple gets sliced in the below
        if row[3]:
            if 'opportunistic invsmnt' in row[3].lower():
                GLOBAL_OPPORTUNISTIC += row[4]

        #Collecting Column 0 and 1
        p1 = row[:2] 
        #Collecting Column 4  --> Used slices in lieu of [4] as this would only return a float
        p2 = row[4:]
        #Converted from tuple to list to allow data manipulation
        new_row = list(p1 + p2)

        #Arrange Strings in array for consistency
        new_row = arrangeStrings(new_row) 
        
        #If there is a value in Location, Location variable will now contain the value 
        if new_row[1]: 
            loc = new_row[1].upper()
            if 'TOTAL' in loc:
                continue  
        
        # example loc = ['TOK', 1684243478.0225258]
        lst.append(loc)
        lst.append(new_row[2])

        #If the first cell in the row[i] contains any value other than None and is not empty
        #Assign the name of that cell to pool      ex(pool = COMPLEX_ALPHA)

        if str(new_row[0]) and str(new_row[0]) != 'None':
            pool = str(new_row[0])

        # example of how data is arranged in system memory 
        #Using a dictionnary---> key :str, value: list[str, float]
        # {'GLOBAL/INTERNATIONAL_EQUITY': ['TOK', 1684243478.0225258]}
        # {'GLOBAL/INTERNATIONAL_EQUITY': ['SING', 2282476822.9744487]}

        if dic.get(pool):
            dic[pool].append(lst.copy()) 
        else: 
            dic[pool] = []
            dic[pool].append(lst.copy())


#List of pools/Buckets
COMPLEX_ALPHA = get_complex_alpha_bos(dic)
CORE_EQUITY = get_core_equity_bos(dic)
FIXED_INCOME = get_fixed_income_bos(dic)
SPECIAL_EQUITY = get_special_equity(dic)
GLOBAL_DOMESTIC = get_global_domestic(dic)
GLOBAL_TOKYO = get_global_tok(dic)
GLOBAL_SINGAPORE = get_global_sing(dic)
GLOBAL_HONGKONG = get_global_hk(dic)

# Start changing width from column C onwards
# param(worksheet, beginning_column, column size)




if calc_isvalid():
    addtofile(COMPLEX_ALPHA, CORE_EQUITY, FIXED_INCOME, SPECIAL_EQUITY,\
    GLOBAL_DOMESTIC, GLOBAL_OPPORTUNISTIC, GLOBAL_TOKYO,\
        GLOBAL_SINGAPORE, GLOBAL_HONGKONG, ws)
    
    print("File is successfully saved")
else:
    print("Error, please check file")
    SystemExit()



changeWidth(ws, 12, 20) #Change width and aplly number-format
    
saveworkbook() 




#print(f"Complex Alpha : {COMPLEX_ALPHA}")
#print(f"Core Equity : {CORE_EQUITY}") 
#print(f"Fixed Income : {FIXED_INCOME}") 
#print(f"Special Equity : {SPECIAL_EQUITY}") 
#print(f"Global Domestic : {GLOBAL_DOMESTIC}")
#print(f"Global Opportunistic  : {GLOBAL_OPPORTUNISTIC}") 
#print(f"Global Tokyo : {GLOBAL_TOKYO}") 
#print(f"Global Singapore : {GLOBAL_SINGAPORE}")       
#print(f"Global Hong Kong : {GLOBAL_HONGKONG}")
#print(f"GRAND_TOTAL : {GRAND_TOTAL}")





    
