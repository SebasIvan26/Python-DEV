################################################################################
##
## BY: Sebastien St Vil
##
## This project can be used freely for all uses, any information in the visual
## interface (GUI) can be modified without any implication.
##
##
################################################################################

import sys as Sys
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
import pyexcel as p
import EIB_AUM as eib_aum

GLOBAL_OPPORTUNISTIC, GRAND_TOTAL = 0, 0
dicNums = {}
grand_Total_position = ''
activateEIB = False 

def xlstoxlsxConverted(sourcepath, destpath):
    p.save_book_as(file_name=sourcepath,\
    dest_file_name=destpath)
    print("Successfully Converted\n")

##Transforms array of Strings into consistent characters and case
def arrangeStrings(array):
    for i in range(len(array)-1):
        if array[i]:
            array[i] = array[i].strip().replace(" ", "_").upper()
    return array 

#Style results range    
def set_border(ws, cell_range):
    #applying border and alignment
        font = Font(size=9)
        align=Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell,'border',border), setattr(cell,'font',font), setattr(cell,'alignment',align)) for cell in flattened]

#Returns the Total amount for the Complex ALpha pool in BOS
def get_complex_alpha_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "COMPLEX_ALPHA" and ("BOS" in array[0] or "RAD" in array[0]): ##Previous prelims used to have a RAD location
                amount = amount + array[1]
    return amount

def get_core_equity_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "CORE_EQUITY" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_fixed_income_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "FIXED_INCOME" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_hk(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "HK" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_london(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "LON" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_sf(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "SF" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_sing(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "SING" in array[0] :
                amount = amount + array[1]
    return amount

def get_global_tok(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "GLOBAL/INTERNATIONAL_EQUITY" and "TOK" in array[0] :
                amount = amount + array[1]
    return amount

def get_special_equity(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if (k == "SPECIAL_EQUITY" or k == "SPECIALTY_EQUITY") and ("BOS" in array[0] or "RAD" in array[0]): ##Previously called "Specialty Equity"
                #Since we might surpass the rows which contain the values, we're ensuring that we're only counting numbers
                if array[1]:
                    amount = amount + array[1]
    return amount

def get_global_domestic(dic):
    print("Testing")
    print(get_global_bos(dic))
    print(get_global_sf(dic))
    print(GLOBAL_OPPORTUNISTIC)
    return get_global_bos(dic) + get_global_sf(dic) - GLOBAL_OPPORTUNISTIC

def calc_isvalid(dic):
    CHECK_TOTAL = get_core_equity_bos(dic) + get_complex_alpha_bos(dic) + get_fixed_income_bos(dic) + get_special_equity(dic) \
        + get_global_domestic(dic) + GLOBAL_OPPORTUNISTIC + get_global_tok(dic) + get_global_sing(dic) + get_global_hk(dic) + get_global_london(dic)
    print('Numbers are being verified....')
    a = CHECK_TOTAL
    b = GRAND_TOTAL
    print(f'Check Total: {a}')
    print(f'Check Total: {b}')
    print(a == b)
    return format(CHECK_TOTAL, '.2f') == format(GRAND_TOTAL, '.2f')

def getStart(ws):
    dataPlacement = 1 ##Variable to determine where data starts, essentially the placement
    for row in ws.iter_rows(min_row=0, max_row=300, min_col=1, max_col=8, values_only=True):
        if str(row[7]) != 'None':
            break;
        dataPlacement += 1
    return dataPlacement + 1

def applyFormat(size,ws):
    column = 1
    while column < 601:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = size
        column += 1

    #Change number to comma-style
    for row in range(1, 400):
        ws["E{}".format(row)].number_format = '#,##0.00_-'
        ws["F{}".format(row)].number_format = '#,##0.00_-'
        ws["G{}".format(row)].number_format = '#,##0.00_-'
        ws["H{}".format(row)].number_format = '#,##0.000_-'
        ws["N{}".format(row)].number_format = '#,##0.00_-'
        ws["O{}".format(row)].number_format = '#,##0.00_-'
   # ecell.number_format = 

def writeToFile(COMPLEX_ALPHA, CORE_EQUITY, FIXED_INCOME, SPECIAL_EQUITY, 
            GLOBAL_DOMESTIC, GLOBAL_OPPORTUNISTIC, GLOBAL_TOKYO,\
            GLOBAL_SINGAPORE, GLOBAL_HONGKONG, GLOBAL_LONDON, ws):
    
    result = {'COMPLEX_ALPHA': COMPLEX_ALPHA, 'CORE_EQUITY': CORE_EQUITY, 'FIXED_INCOME': FIXED_INCOME, 'SPECIAL_EQUITY': SPECIAL_EQUITY,
    'GLOBAL_INTL_EQUITY_DOMESTIC': GLOBAL_DOMESTIC, 'GLOBAL_INTL_EQUITY_OPPORTUNISTIC': GLOBAL_OPPORTUNISTIC, 'GLOBAL INTL_EQUITY_TOKYO': GLOBAL_TOKYO,\
        'GLOBAL_INTL_EQUITY_SINGAPORE': GLOBAL_SINGAPORE, 'GLOBAL_INTL_EQUITY_HONGKONG':GLOBAL_HONGKONG, 'GLOBAL_INTL_EQUITY_LONDON':GLOBAL_LONDON}
    
    global dicNums
    global grand_Total_position
    total = 0
    starting_row, starting_col = 8, 12
    for k, v in result.items():
        total += v
        if 'domestic' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Domestic'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'opportunistic' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Opportunistic'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'tokyo' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Tokyo'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'singapore' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Singapore'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'hong' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'Hong Kong'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        elif 'london' in k.lower():
            ws.cell(row=starting_row, column=starting_col).value = k[:18]
            ws.cell(row=starting_row, column=starting_col + 1).value = 'London'
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
        else:
            ws.cell(row=starting_row, column=starting_col).value = k
            ws.cell(row=starting_row, column=starting_col + 2).value = v
            starting_row += 1
    
    print("File is being generated......")
    
    #At this stage, this is pointing to the end row
    starting_row += 2
    ws.cell(row=starting_row, column=starting_col).value = 'Total'

    for i in range(8,18):
        co = 'O' + str(i)
        cn = 'N' + str(i)
        ws[co] ='=ROUND('+cn+'/1000,2)'


    applyFormat(23,ws)
    ws['N19'] = '=SUM(N8:N17)'
    ws['O19'] = '=SUM(O8:O17)'
    ws['N20'] = grand_Total_position
    ws['N21'] = '=N19-N20'
    ws['N21'].style = 'Comma'
    #Apply custom format to cells (fix cells' lengths and number formats)
    
    set_border(ws, 'L8:O21')

    if not calc_isvalid(dicNums):
        ws['L23'] = 'Error: Calculations are not valid'
        ws['L23'].style = 'Warning Text'

def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving file")
    
    print("File is successfully saved in specified file path.")

#Load existing spreadsheet
def checkAndLoad(source):
    #Create active worksheet
    try:
        wb = load_workbook(source)
        ws = wb.active
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")

    try:
        return [wb, ws]
    except Exception:
        print("Error in returning workbook")


def processFromSheet(ws):
    #We use a dictionnary to chain the key(pool Bucket) to values
    dic = {}
    global GLOBAL_OPPORTUNISTIC
    global GRAND_TOTAL
    global grand_Total_position
    row_loc = 8
    loc, pool = '', ''

    for row in ws.iter_rows(min_row=8, max_row=300, min_col=1, max_col=5, values_only=True):
        if row[0]:
            if 'grand total' in row[0].lower():
                GRAND_TOTAL = 0
                GRAND_TOTAL += row[4]
                grand_Total_position = "=E" + str(row_loc)
        if 'Total'.upper() not in str(row[0]).upper(): # as long as the first row does not contain the word "Total" 
            #Upper is used for case Insensitive Comparison)
            lst = []

            #Get Opportunistic before "Account Names" are unavailable when tuple gets sliced in the below
            if row[3]:
                op = row[3].lower()
                if 'opportunistic invstmnt' in op or 'opportunistic invsmnt'in op: #Account for the fact that Opportunistic is written inconistently in prelim
                    GLOBAL_OPPORTUNISTIC += row[4]
                    print("Print Global_opportunistic")
                    print(GLOBAL_OPPORTUNISTIC)

            #Collecting Column 0 and 1
            p1 = row[:2] 
            #Collecting Column 4  --> Used slices in lieu of [4] as this would only return a float
            p2 = row[4:]
            #Converted from tuple to list to allow data manipulation
            new_row = list(p1 + p2)

            #Arrange Strings in array for consistency
            new_row = arrangeStrings(new_row) 
            
            #If there is a value in Location, Location variable will now contain the value 
            if new_row[1] and 'BLANK' not in new_row[1].upper(): 
                loc = new_row[1].upper()
                if 'TOTAL' in loc:
                    continue  

            # example loc = ['TOK', 1684243478.0225258]
            lst.append(loc)
            lst.append(new_row[2])

            #If the first cell in the row[i] contains any value other than None and is not empty
            #Assign the name of that cell to pool      ex(pool = COMPLEX_ALPHA)

            if str(new_row[0]) and str(new_row[0]) != 'None':
                pool = str(new_row[0])

            # example of how data is arranged in system memory 
            #Using a dictionnary---> key :str, value: list[str, float]
            # {'GLOBAL/INTERNATIONAL_EQUITY': ['TOK', 1684243478.0225258]}
            # {'GLOBAL/INTERNATIONAL_EQUITY': ['SING', 2282476822.9744487]}

            if dic.get(pool):
                dic[pool].append(lst.copy()) 
            else: 
                dic[pool] = []
                dic[pool].append(lst.copy())

        row_loc += 1

    return dic


def calculateFromDic(dic,ws):
    global dicNums
    dicNums = dic.copy()
    print('Verified....')
    
    COMPLEX_ALPHA = get_complex_alpha_bos(dic)
    CORE_EQUITY = get_core_equity_bos(dic)
    FIXED_INCOME = get_fixed_income_bos(dic)
    SPECIAL_EQUITY = get_special_equity(dic)
    GLOBAL_DOMESTIC = get_global_domestic(dic)
    GLOBAL_TOKYO = get_global_tok(dic)
    GLOBAL_SINGAPORE = get_global_sing(dic)
    GLOBAL_HONGKONG = get_global_hk(dic)
    GLOBAL_LONDON = get_global_london(dic)

            
    writeToFile(COMPLEX_ALPHA, CORE_EQUITY, FIXED_INCOME, SPECIAL_EQUITY,\
    GLOBAL_DOMESTIC, GLOBAL_OPPORTUNISTIC, GLOBAL_TOKYO,\
    GLOBAL_SINGAPORE, GLOBAL_HONGKONG, GLOBAL_LONDON, ws)


def main(bucketSourcePath, bucketDestPath, ACTIVATE_EIB):
    global GLOBAL_OPPORTUNISTIC
    #########################SOURCE FILE LOCATION#################################################
    source = bucketSourcePath

    #########################DESTINATION FILE LOCATION#################################################
    destination = bucketDestPath +'.xlsx' if '.xlsx' not in bucketDestPath else bucketDestPath

    #XLS to XLSX Converter - If file is Excel 97-2003 default:Disactivated
    #Turn True to activate
    xlstoxlsxConverted(source, destination)

    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(destination)

    try:
        wb = wbws[0]
        ws = wbws[1]
    except Exception:
        print("Error in loading file....")

    #INSERT EXCEL NUMBERS INTO DATA STRUCTURES TO ORGANIZE (ORGANIZED BY POOLS)
    dic = processFromSheet(ws)

    #CALCULATE BUCKETS WITH DATA FROM DICTIONNARY
    #CHECK IF CALCULATIONS ARE CORRECT (SUM OF NUMBERS AMASSED IN ABOVE FUNCTION IS COMPARED TO GRAND TOTAL)--> THIS FUNCTION RETURNS
    #TRUE IF THEY ARE EQUAL OTHERWISE FALSE
    calculateFromDic(dic,ws)

    #CHECK FILE IS SAVED PROPERLY
    checkAndSave(wb, destination)


##################This activates only if EIB box is checked #######################
    if ACTIVATE_EIB:
        eib_aum.main(dic, bucketDestPath)
    else:
        pass

    ##Rest Global Variable so it resets as programs is still at runtime
    GLOBAL_OPPORTUNISTIC = 0  #<--------Inserted at end of program so that it resets without impacting eib_aum


    #######Returning dic in order for STAT AUM data to be accessible for EIB Generation
    return dic

if __name__ == "__main__":
    main(bucketSourcePath, bucketDestPath, ACTIVATE_EIB)





    
