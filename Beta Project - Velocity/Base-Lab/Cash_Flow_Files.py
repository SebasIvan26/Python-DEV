################################################################################
##
## BY: Sebastien St Vil
##
## This project can be used freely for all uses, any information in the visual
## interface (GUI) can be modified without any implication.
##
##
################################################################################

import platform
import sys as Sys
import pandas as pd
import numpy as np
import matplotlib
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


cellTop, cellBottom = 0, 0
grand_total_pos = 11 #Number chosen so that row corresponds to sheet
Quarter, pTabName = '', ''
strtdate, dataTab = '', 'Memo'
boiteDeCrayon = ['7FD5FF','2F8B1A','F8FA4E','4E64D2','000000','F98107','F90707','FDB4B4', '3CA463'] 
                 #0:bleupale, 1:vert, 2:jaune, 3:bleu, 4:noir, 5:jaune abricot, 6:rouge, 7:rose, 8:lightgreen


def getColorFill(couleur):
    blueFill = PatternFill(start_color=couleur,
                   end_color=couleur,
                   fill_type='solid')
    return blueFill

def applyFormat(size,ws):
    column = 1
    while column < 601:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = size
        column += 1

    #Change number to comma-style
    for row in range(1, 400):
        ws["H{}".format(row)].number_format = '#,##0.00_);[Red](#,##0.00)'
        ws["I{}".format(row)].number_format = '#,##0.00_);[Red](#,##0.00)'
        ws["J{}".format(row)].number_format = '#,##0.00_);[Red](#,##0.00)'
   # ecell.number_format = 

def applyFormat3(ws, rowtop, rowbottom, colstart, colend):
    global boiteDeCrayon

    for rows in ws.iter_rows(min_row=rowtop, max_row=rowbottom, min_col=colstart, max_col=colend):
        for cell in rows:
            cell.fill = getColorFill(boiteDeCrayon[7])
            cell.number_format = '#,##0.00'   

def checkAndSave(wb,dest):
    wb.save(dest)
    wb.close()
    

def duplicateWB(wb, ws):
    global dataTab
    source = ws
    target = wb.copy_worksheet(source)
    target.title = dataTab
    target.sheet_properties.tabColor = 'FF0000'
    return target

#Load existing spreadsheet
def checkAndLoad(source):
    global pTabName
    #Create active worksheet
    try:
        #wb = load_workbook(filename=source, read_only=True)
        wb = load_workbook(filename=source, data_only=True)
        ws = wb.sheetnames
        
        #Ensuring to load the "WRPP" workbook
        for s in ws:
            if 'pivot'.upper() in s.upper():
                pivottabname = s
                pTabName = pivottabname
            elif 'detail'.upper() in s.upper():
                detailtabname = s
                #Selecting Worksheets
        pivottab = wb[pivottabname]
        detailtab = wb[detailtabname]
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")
    
    try:
        return [wb, pivottab, detailtab]
    except Exception:
        print("Error in returning workbook")



#Load updated Worksheet and Apply Formatting
def checkAndLoad2(wb, source):
    global boiteDeCrayon
    #Create active worksheet
    try:
        #wb = load_workbook(filename=source, read_only=True)
        #wb = load_workbook(filename=source, data_only=True)
        wslist = wb.sheetnames
        print(f"Sheets in WorkBook after Pandas: {wslist}")
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Updated file loaded...")
    
    #print("Sheet in workbook after pandas")
    #print(wb.sheetnames)
    stattab1 = wb['NB']
    stattab2 = wb['LB']
    stattab3 = wb['OCF']
    
    stattab1.sheet_properties.tabColor = boiteDeCrayon[4]
    stattab2.sheet_properties.tabColor = boiteDeCrayon[4]
    stattab3.sheet_properties.tabColor = boiteDeCrayon[4]

    column = 1
    while column < 60:
        i = get_column_letter(column)
        stattab1.column_dimensions[i].width = 35
        stattab2.column_dimensions[i].width = 35
        stattab3.column_dimensions[i].width = 35
        column += 1


    for i in range(2, 6):
        stattab1.cell(row=i,column=2).fill = getColorFill(boiteDeCrayon[8])
        stattab2.cell(row=i,column=2).fill = getColorFill(boiteDeCrayon[7])
        stattab3.cell(row=i,column=2).fill = getColorFill(boiteDeCrayon[2])
    
    try:
        wb.save(source)
    except Exception:
        print("File unable to be saved")

def getCellPos(ws):
    global Quarter
    global cellTop
    global cellBottom
    global grand_total_pos
    
    for row in ws.iter_rows(min_row=11, max_row=700, min_col=1, max_col=5, values_only=True):
        if row[0]:
            if 'grand total' in str(row[0]).lower():
                print(f"Location position is {grand_total_pos}")
                break
        grand_total_pos += 1

    if Quarter == 1:
        cellTop = grand_total_pos - 15
        cellBottom = grand_total_pos - 1
    elif Quarter == 2:
        cellTop = grand_total_pos - 18
        cellBottom = grand_total_pos - 1
    elif Quarter == 3:
        cellTop = grand_total_pos - 21
        cellBottom = grand_total_pos - 1
    else:
        cellTop = grand_total_pos - 24
        cellBottom = grand_total_pos - 1
    
    getstrtdate(ws)
    print(f'Cell top is {cellTop}')
    print(f'Cell bottom is {cellBottom}')

def getstrtdate(ws):
    global cellTop
    global strtdate

    pos = 'A' + str(cellTop)
    strtdate = str(ws[pos].value).strip()

def getQuarter(ws):
    global Quarter

    #Gets the date in Cell A2 to identify Quarter
    date = ws['A2'].value.strip()[27:32]

    period = ['06/30', '09/30', '12/31', '03/31']

    #Loops through period in order to determine corresponding Quarter
    for i, e in enumerate(period):
        if date in e:
            Quarter = i + 1
            break
    print(f'The Quarter is {Quarter}')

    if not Quarter: #Clause to identify whether Quarter was identified
        print("Quarter was not properly identified")

def processWorksheet(ws):
    global Quarter
    global cellTop
    global cellBottom
    global grand_total_pos
    global boiteDeCrayon


    #Reminder to put in correct BPS for Quarter 4
    qBPS = {1: [0.5, 1.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 1.5, 0.5], 
            2: [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 5.5, 5.5, 5.5, 5.5, 5.5, 5.5, 5.5, 4.5, 3.5, 2.5, 1.5, 0.5],
            3: [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5, 7.5, 8.5, 8.5, 8.5, 8.5, 8.5, 7.5, 6.5, 5.5, 4.5, 3.5, 2.5, 1.5, 0.5],
            4: [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5, 7.5, 8.5, 9.5, 10.5, 11.5, 11.5, 10.5, 9.5, 8.5, 7.5, 6.5, 5.5, 4.5, 3.5, 2.5, 1.5, 0.5]}
    currentBPS = qBPS[Quarter]
    print(currentBPS)
    NBcol, LBcol, ECFcol, BPScol = 8, 9, 10, 12
    
    for i in range(cellTop, cellBottom + 1):
        rowNum = str(i)
        ws.cell(row=i, column=NBcol).value = '=B'+ rowNum + '/12*L' + rowNum 
        ws.cell(row=i, column=LBcol).value = '=C'+ rowNum + '/12*L' + rowNum 
        ws.cell(row=i, column=ECFcol).value = '=D'+ rowNum + '/12*L' + rowNum  
        ws.cell(row=i, column=BPScol).value = currentBPS[i - cellTop]
        ws.cell(row=i, column=NBcol).fill = getColorFill(boiteDeCrayon[0])
        ws.cell(row=i, column=LBcol).fill = getColorFill(boiteDeCrayon[0])
        ws.cell(row=i, column=ECFcol).fill = getColorFill(boiteDeCrayon[0])

    
    print("After Loop")
    print(f'Celltop is now {cellTop}')

    ws.cell(row=cellTop-1, column=NBcol).value = 'NB'
    ws.cell(row=cellTop-1, column=NBcol).fill = getColorFill(boiteDeCrayon[0])

    ws.cell(row=cellTop-1, column=LBcol).value = 'LB'
    ws.cell(row=cellTop-1, column=LBcol).fill = getColorFill(boiteDeCrayon[0])

    ws.cell(row=cellTop-1, column=ECFcol).value = 'ECF'
    ws.cell(row=cellTop-1, column=ECFcol).fill = getColorFill(boiteDeCrayon[0])

    ws.cell(row=grand_total_pos, column=NBcol).value = '=SUM(H' +str(cellTop) + ':H' + str(cellBottom) + ')'
    ws.cell(row=grand_total_pos, column=NBcol).fill = getColorFill(boiteDeCrayon[0])

    ws.cell(row=grand_total_pos, column=LBcol).value = '=SUM(I' +str(cellTop) + ':I' + str(cellBottom) + ')'
    ws.cell(row=grand_total_pos, column=LBcol).fill = getColorFill(boiteDeCrayon[0])


    ws.cell(row=grand_total_pos, column=ECFcol).value = '=SUM(J' +str(cellTop) + ':J' + str(cellBottom) + ')'
    ws.cell(row=grand_total_pos, column=ECFcol).fill = getColorFill(boiteDeCrayon[0])

    applyFormat(23, ws)
    ws.sheet_properties.tabColor = "289C0E" 

#Insert the 3 columns before column labeled "Product"
def getColPosition(ws, mot):
    insertColPos = 0

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=21, values_only=True):
        if col[0]:
            print(str(col[0]).lower())
            wd = str(col[0]).strip().lower()
            if mot in wd:
                print(f"{mot} is found in position {insertColPos}")
                break
        insertColPos += 1
    return insertColPos + 1

def getPeriodStart(ws):
    global strtdate

    print(f"Global start date is {strtdate}")

    starting_row = 0
    maxrow = ws.max_row
    print(f"Maxrow is {maxrow}")
    for row in ws.iter_rows(min_row=1, max_row=maxrow, min_col=1, max_col=21, values_only=True):
        if row[0]:
            if strtdate in str(row[0]):
                print(f'Row 0 is {str(row[0])}')
                break
        starting_row += 1
    return starting_row + 1
    
def processWorksheet2(ws):
    global pTabName
    global cellTop
    global cellBottom

    
    insertColPos = getColPosition(ws, 'product') #Get position of column 'Product' and insert 3 cols before it
    ws.insert_cols(insertColPos, 3)

    #Obtain column Letter of 3 added columns
    col1 = get_column_letter(insertColPos) 
    col2 = get_column_letter(insertColPos + 1) 
    col3 = get_column_letter(insertColPos + 2) 

    #Title column name
    ws[col1 +'1'], ws[col2+'1'], ws[col3+'1'] = 'NB', 'LB', 'OCF'

    pstart = getPeriodStart(ws)

    print(f"The starting position of the current period is {pstart}")

    #Inner Function returns position in # format, outer converts to letter
    nb_ABR_Pos = getColPosition(ws, 'abr')
    index1, index2, index3 = get_column_letter(nb_ABR_Pos), get_column_letter(nb_ABR_Pos + 1), get_column_letter(nb_ABR_Pos + 2)
    print("NBR Position")
    print(index1)
    print(pstart)

    for i in range(pstart, ws.max_row):
        j = str(i)
        top = str(cellTop)
        bottom = str(cellBottom)
        a = '=' + index1 + j + '/12*XLOOKUP(@A:A,' + pTabName + '!$A$' + top +\
             ':$A$' + bottom + ',' + pTabName + '!$L$' + top + ':$L$' + bottom + ',1,0)'

        b = '=' + index2 + j + '/12*XLOOKUP(@A:A,' + pTabName + '!$A$' + top +\
             ':$A$' + bottom + ',' + pTabName + '!$L$' + top + ':$L$' + bottom + ',1,0)'
        
        c = '=' + index3 + j + '/12*XLOOKUP(@A:A,' + pTabName + '!$A$' + top +\
             ':$A$' + bottom + ',' + pTabName + '!$L$' + top + ':$L$' + bottom + ',1,0)'

        ws[col1+j] = a
        ws[col2+j] = b
        ws[col3+j] = c

    #Get sum     
    maxim = str(ws.max_row-1)
    maxim1 = str(ws.max_row)
    ws[col1+maxim1] = '=SUM(' + col1 + str(pstart) + ':' + col1 + maxim + ')' 
    ws[col2+maxim1] = '=SUM(' + col2 + str(pstart) + ':' + col2 + maxim + ')' 
    ws[col3+maxim1] = '=SUM(' + col3 + str(pstart) + ':' + col3 + maxim + ')' 

    ##Aplly Formatting
    applyFormat(23, ws)
    applyFormat3(ws, pstart, int(maxim1), insertColPos, insertColPos+2)

def writeStatsToFile(nbStat, lbStat, ocfStat, destination):
    try:
        writer = pd.ExcelWriter(destination, engine='openpyxl', mode='a')
        writer.book = load_workbook(destination)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        nbStat.to_excel(writer, sheet_name='NB', freeze_panes=(1,0))
        lbStat.to_excel(writer, sheet_name='LB', freeze_panes=(1,0))
        ocfStat.to_excel(writer, sheet_name='OCF', freeze_panes=(1,0))
        writer.save()
        writer.close()
    except Exception:
        print("Statistic Performance Tabs are not properly saved")

    #############    ####################################
    #checkAndSave(wb, "tgjrigjsdf" + destination)
    ##################When exception is created due to file not existing, file looks appropriate

    #checkAndSave(wb, destination)

def getDataStat(filepath, memoTName):
    
    memo_data_df = pd.read_excel(filepath, sheet_name=memoTName, usecols=['Account Name', 'NB', 'LB', 'OCF'])
    memo_data_df = memo_data_df.dropna()
    df_source = memo_data_df

    df1 = df_source.groupby(['Account Name'])['NB'].sum().sort_values(ascending=False)
    df1.loc['Total'] = df1.sum()

    df2 = df_source.groupby(['Account Name'])['LB'].sum().sort_values(ascending=True)
    df2.loc['Total'] = df2.sum()

    df3 = df_source.groupby(['Account Name'])['OCF'].sum().sort_values(ascending=True)
    df3.loc['Total'] = df3.sum()

    return [df1, df2, df3]

def main():
    system = platform.system() ##Windows or MAC
    global dataTab
    #########################SOURCE FILE LOCATION#################################################
    source = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/wtc_cash_flows/WTC Cash Flow Raw.xlsx'\
            if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\110-Trust (WTC)\Internal Reporting\Board Reports\2021\Q3\Support\testing.xlsx'

    #########################DESTINATION FILE LOCATION#################################################
    #destination = bucketDestPath +'.xlsx' if '.xlsx' not in bucketDestPath else bucketDestPath


    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(source)

    try:
        wb = wbws[0]
        pivottab = wbws[1]
        detailtab = wbws[2]
    except Exception:
        print("Error in loading file....")

    #Identifies the Quarter which the Cash FLow Files is based off    
    getQuarter(pivottab)

    #Determine cell positions for formulas
    getCellPos(pivottab)

    #Process Pivot Tab
    processWorksheet(pivottab)

    #Create copy of Detail Tab in roder to leave original data intact 
    memoTab = duplicateWB(wb, detailtab)

    processWorksheet2(memoTab)


    destination = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/wtc_cash_flows/WTC Cash Flow testingRes.xlsx'\
            if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\110-Trust (WTC)\Internal Reporting\Board Reports\2021\Q3\Support\testingres.xlsx'

    

    #######CHECK FILE IS SAVED PROPERLY#########
    checkAndSave(wb, destination)

    
    ##Read and Analyze file with Pandas
    stats_pivot_tables = getDataStat(destination, dataTab)

    nbStat = stats_pivot_tables[0]
    lbStat = stats_pivot_tables[1]
    ocfStat = stats_pivot_tables[2]

    print(nbStat)
    print(lbStat)
    print(ocfStat)

    checkAndSave(wb, destination)

    #writeStatsToFile(nbStat, lbStat, ocfStat, wb2, destination)


    #Load Updated File and Save
    #checkAndLoad2(wb2, destination)

    #processWorksheet(pivottab)

    #checkAndSave(wb2, destination)

    #writeStatsToFile(nbStat, lbStat, ocfStat, destination)

if __name__ == "__main__":
    main()





    
