import sys as Sys
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
import xlrd
import EIB_REV as eib_rev

WRPP_SHEET_NAME = ""
dicNums = {} #Holds the results for the pools, noot the formulas 

def xlstoxlsxConverted(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()
    try:
        sheet_names = book_xls.sheet_names()[:2] #Only returns the first 2 tabs which is all that we need
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

        book_xlsx.save(dst_file_path)
    except Exception:
        print("Error in conversion")
    print("Successfully Converted\n")

##Transforms Strings into consistent characters and case
def arrangeStrings(array):
    array = array.strip().replace(" ", "_").upper()
    return array 

def set_border(ws, cell_range):
    #applying border and alignment
        font = Font(size=9)
        align=Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell,'border',border), setattr(cell,'font',font), setattr(cell,'alignment',align)) for cell in flattened]

def writeToFile(dic, ws):
    total = 0
    starting_row, starting_col = 8, 12
    for k, v in dic.items():
        if "GRAND" in k:
            continue
        ws.cell(row=starting_row, column=starting_col).value = k
        ws.cell(row=starting_row, column=starting_col + 1).value = v
        starting_row += 1
    
    print("File is being generated......")
    
    #At this stage, this is pointing to the end row

    ws.cell(row=starting_row + 1, column=starting_col).value = 'Total'

    for i in range(8,13):
        co = 'N' + str(i)
        cn = 'M' + str(i)
        ws[co] ='=ROUND('+cn+'/1000,2)'

##Excel checks
    ws['M14'] = '=SUM(M8:M12)'
    ws['N14'] = '=SUM(N8:N12)'
    ws['M15'] = dic['GRAND_TOTAL']
    ws['M16'] = '=M14-M15'
    ws['M16'].style = 'Comma'

    #Set border
    set_border(ws, 'L8:N16')

    #Check to let user know if CAlc is invalid
    if not calc_isvalid():
        ws['L19'] = 'Error: Calculations are not valid'
        ws['L19'].style = 'Warning Text'



def calc_isvalid():
    global dicNums
    pool_Totals = 0
    for k, v in dicNums.items():
        if 'GRAND' in k:
            continue
        pool_Totals += v 
    print(f'Pool Totals: {pool_Totals}')
    print(dicNums['GRAND_TOTAL'])
    return format(pool_Totals, '.3f') == format(dicNums['GRAND_TOTAL'], '.3f')


def getStart(ws):
    dataPlacement = 1 ##Variable to determine where data starts, essentially the placement
    for row in ws.iter_rows(min_row=0, max_row=300, min_col=1, max_col=8, values_only=True):
        if str(row[6]) != 'None':
            break;
        dataPlacement += 1
    return dataPlacement + 1

def applyFormat1(size,ws):
    column = 1
    while column < 601:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = size
        column += 1

    #Change number to comma-style
    for row in range(1, 400):
        ws["E{}".format(row)].number_format = '#,##0.00_-'
        ws["F{}".format(row)].number_format = '#,##0.00_-'
        ws["G{}".format(row)].number_format = '#,##0.00_-'
        ws["H{}".format(row)].number_format = '#,##0.000_-'
        ws["I{}".format(row)].number_format = '#,##0.00_-'
        ws["J{}".format(row)].number_format = '#,##0.00_-'
        ws["L{}".format(row)].number_format = '#,##0.00_-'
        ws["M{}".format(row)].number_format = '#,##0.00_-'
        ws["N{}".format(row)].number_format = '#,##0.00_-'
        ws["O{}".format(row)].number_format = '#,##0.00_-'

def applyFormat2(size,ws):
    column = 1
    while column < 601:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = size
        column += 1

    #Change number to comma-style
    for row in range(1, 100):
        ws["C{}".format(row)].number_format = '#,##0.00_-'
        ws["F{}".format(row)].number_format = '#,##0.00_-'
        ws["G{}".format(row)].number_format = '#,##0.00_-'
        ws["H{}".format(row)].number_format = '#,##0.00_-'

#Load existing spreadsheet
def checkAndLoad(source):
    global WRPP_SHEET_NAME
    #Create active worksheet
    try:
        wb = load_workbook(source)
        ws = wb.sheetnames
        
        #Ensuring to load the "WRPP" workbook
        for s in ws:
            if 'bucket'.upper() not in s.upper() and 'WRPP'.upper() in s.upper():
                wrpp = s
            elif 'BUCKET REPORT' in s.upper():
                ex_wrpp = s
                #Selecting WRPP Worksheet
        ws = wb[ex_wrpp]
        ws2 = wb[wrpp]
        WRPP_SHEET_NAME = wrpp + '!'
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")
    
    try:
        return [wb, ws, ws2]
    except Exception:
        print("Error in returning workbook")

def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving file")
    
    print("File is successfully saved in specified file path.")

def processFromSheet2(ws):
    global dicNums
    applyFormat2(20, ws)
    ws.sheet_properties.tabColor = 'F98107'
    
    #Dictionnary to store 'pool name totals' as key and 'Monthly Base Revenue' position's location as values
    dic = {}
    row_loc = 8
    for row in ws.iter_rows(min_row=row_loc, max_row=50, min_col=1, max_col=7, values_only=True):
        if row[0]:
            val = arrangeStrings(str(row[0]))
            if 'Total'.upper() in val:
                dicNums[val] = row[6] #Storing values (not formulas)
                dic[val] = '=' + WRPP_SHEET_NAME + 'G'+ str(row_loc)
        row_loc += 1
    #ex: dic[COMPLEX ALPHA TOTAL] = WRPP!G9
    return dic

def processFromSheet1(ws, dic):
    global dicNums
    applyFormat1(20, ws)
    ws.sheet_properties.tabColor = '2F8B1A'
    row_loc = 8 #Get the row at which data starts
    dicto = {}
    col1, col2 = 9, 10
    for row in ws.iter_rows(min_row=8, max_row=300, min_col=1, max_col=7, values_only=True):
        if row[0]:
            val = arrangeStrings(str(row[0]))
            if val in dic:
                dicNums[val] += row[6]  #Storing values (not formulas)
                #Link Bucket pool totals to their respective pools in WRPP
                ws.cell(row=row_loc, column = col1).value = dic[val]
                ws.cell(row=row_loc, column = col2).value = '=' + get_column_letter(col1) + str(row_loc) + '+' + 'G' + str(row_loc)
                dicto[val] = '=' + 'J' + str(row_loc)
                #Get cell references for Final 
        row_loc += 1
    return dicto #This returns references to results in Column J

def main(bucketSourcePath, bucketDestPath, ACTIVATE_EIB):

    #########################SOURCE FILE LOCATION#################################################
    source = bucketSourcePath

    #########################DESTINATION FILE LOCATION#################################################
    destination = bucketDestPath +'.xlsx' if '.xlsx' not in bucketDestPath else bucketDestPath

    #XLS to XLSX Converter - If file is Excel 97-2003 default:Disactivated
    #Turn True to activate
    xlstoxlsxConverted(source, destination)

    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(destination)

    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(destination)

    try:
        wb = wbws[0]
        ws1 = wbws[1]
        ws2 = wbws[2]
    except Exception:
        print("Error in processing file....")

    #INSERT EXCEL NUMBERS INTO DATA STRUCTURES TO ORGANIZE (ORGANIZED BY POOLS)
    dic = processFromSheet2(ws2)
    dicto = processFromSheet1(ws1, dic)
    writeToFile(dicto, ws1)
    checkAndSave(wb, destination)
    

    #######Returning dic in order for STAT REV data to be accessible for EIB Generation
    if ACTIVATE_EIB:
        eib_rev.main(dicNums, bucketDestPath)
    else:
        pass

    return dicNums

if __name__ == "__main__":
    main(bucketSourcePath, bucketDestPath, ACTIVATE_EIB)