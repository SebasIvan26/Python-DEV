import sys as Sys
import requests
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from datetime import date, timedelta
from STAT_AUM import main as main2

checktotal = 0
debit_Col = 9
credit_Col = 10

def getPrevMonthEndDate():
    #ex: 2021-01-31 **Changes to 1/31/21 when inserted into excel
    last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
    #start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
    return last_day_of_prev_month

def getMonthYear():
    #ex: February 2021 --> str
    prev_month_year = date.today().replace(day=1) - timedelta(days=1)
    prev_month_year = prev_month_year.strftime("%d %B, %Y")
    prev_month_year = prev_month_year[3:]
    prev_month_year = prev_month_year.replace(',', '')
    return prev_month_year

def applyFormat(ws):

    #Change number to comma-style
    for row in range(1, 100):
        ws["N{}".format(row)].number_format = "MM/DD/YYYY"


def getStart(ws):
    dataPlacement = 1 ##Variable to determine where data starts, essentially the placement
    for row in ws.iter_rows(min_row=0, max_row=300, min_col=1, max_col=8, values_only=True):
        if 'BU' in str(row[4]):
            break;
    return dataPlacement + 1

def updateImportAccountingTab(ws):
    row_loc = 6

    for row in ws.iter_rows(min_row=row_loc, max_row=21, min_col=1, max_col=23, values_only=True):
        ws.cell(row=row_loc, column=14).value = getPrevMonthEndDate()
        ws.cell(row=row_loc, column=16).value = getMonthYear() + " STAT AUM"
        row_loc += 1

def getAUMTemplate(source):
    url = "https://github.com/SebasIvan26/templates/blob/master/AUM%20EIB.xlsx?raw=true"

    username = 'kseb0@hotmail.com'
    token =  '538cd0c05c0207b17a43ca6324f12191a73bb696'

    r = requests.get(url,auth=(username,token))
    path = source
    with open(path, 'wb') as f:
        f.write(r.content)
    return path

def get_fixed_income_bos(dic):
    amount = 0
    for k, v in dic.items():
        for array in v:
            if k == "FIXED_INCOME" and "BOS" in array[0] :
                amount = amount + array[1]
    return amount

#Load existing spreadsheet
def checkAndLoad(source):
    #Create active worksheet
    getAUMTemplate(source)
    try:
        wb = load_workbook(source)
        ws = wb.sheetnames
        
        #Ensuring to load the "WRPP" workbook
        for s in ws:
            if 'Import Account'.upper() in s.upper():
                import_tab = s
            elif 'Journal Entry'.upper() in s.upper():
                entry_tab = s
                #Selecting WRPP Worksheet
        ws = wb[import_tab]
        ws2 = wb[entry_tab]
    #except openpyxl.utils.exceptions.InvalidFileException:
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")
    
    return [wb, ws, ws2]

def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving file")
    
    print("File is successfully saved in specified file path.")

# {'GLOBAL/INTERNATIONAL_EQUITY': ['TOK', 1684243478.0225258]}
def processFromSheet2(ws):
    global credit_Col
    global dicNums

    #print(GRAND_TOTAL)
    row_loc = 5
    for row in ws.iter_rows(min_row=row_loc, max_row=50, min_col=1, max_col=46, values_only=True):
        if row[4]:
            if(row_loc == 6 and str(row[5]) == 'Ledger Account'):
                ws.cell(row= row_loc, column=debit_Col).value = 0
            if(str(row[5]) == 'Statistical AUM'):
                ws['K7'].value = '=Test'
                print(row[2])
        row_loc += 1

def main():

    #########################SOURCE FILE LOCATION#################################################

    source = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Custom - GIT pulls copy/Base-master/EIBTemplate.xlsx'

    #########################DESTINATION FILE LOCATION#################################################
    destination = source

    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(source)

    wb = wbws[0]
    ws1 = wbws[1]
    ws2 = wbws[2]


    #INSERT EXCEL NUMBERS INTO DATA STRUCTURES TO ORGANIZE (ORGANIZED BY POOLS)
    updateImportAccountingTab(ws1)
    processFromSheet2(ws2)
    applyFormat(ws1)
    #dicto = processFromSheet1(ws1, dic)
    checkAndSave(wb, destination)


if __name__ == "__main__":
    bucketSourcePath = '/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Custom - GIT pulls copy/Base-master/Bucket Report(raw).xlsx'
    bucketDestPath = '/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Custom - GIT pulls copy/Base-master/bucketRes'
    #checktotal = main2(bucketSourcePath, bucketDestPath)
    #print(f'{checktotal} is EIB AUM')
    main()
    