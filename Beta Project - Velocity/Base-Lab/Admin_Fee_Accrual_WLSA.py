import sys as Sys
import platform
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from datetime import date, timedelta
from difflib import get_close_matches

spend_category_dic = {}

########Load Data Audit File########
def checkAndLoad_Data_Audit(path):
    try:
        wb = load_workbook(path)
        ws = wb.active #active means last opened sheet
    except Exception:
        print("Unable to load Data Audit File")
        return
    return ws

########Load WMF Expense Accrual File##########
def checkAndLoad_expense(source):
    #Create active worksheet
    try:
        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb.sheetnames
        
        #Ensuring to load the Workbook and "By Spend Category For Accrual" worksheet
        for s in ws:
            if 'spend'.upper() in s.upper() and 'accrual'.upper() in s.upper():
                accrual = s
                print(accrual)
                break
        ws = wb[accrual]
    except Exception:
        print("Unable to load file Admin Fee file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Admin Fee Expense Accrual loaded...")
    try:
        return ws
    except Exception:
        print("Error in returning workbook")
#################Saves Admin Fee EIB######################
def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving Admin Fee EIB file")
    
    print("EIB File is successfully saved in specified file path.")

#############################Loads Admin Fee EIB###########################
def checkAndLoad_EIB(source):
    #Create active worksheet
    #getWebAUMTemplate(source) Downloads Template from github repository -->Disactivated
    try:
        wb = load_workbook(source)
        ws = wb.sheetnames
        
        #Ensuring to load the "WRPP" workbook
        for s in ws:
            if 'Import Account'.upper() in s.upper():
                import_tab = s
            elif 'Journal Entry'.upper() in s.upper():
                entry_tab = s
                #Selecting WRPP Worksheet
        ws = wb[import_tab]
        ws2 = wb[entry_tab]
    #except openpyxl.utils.exceptions.InvalidFileException:
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")
    
    return [wb, ws, ws2]

################Function is used to obtain relevant date to insert in EIB#######################
def getPrevMonthEndDate():
    #ex: 2021-01-31 **Changes to 1/31/21 when inserted into excel
    last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
    #start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
    return last_day_of_prev_month

################Function is used to obtain relevant date to insert in EIB#######################
def getPrevMonthYear():
    #ex: February 2021 --> str
    prev_month_year = date.today().replace(day=1) - timedelta(days=1)
    prev_month_year = prev_month_year.strftime("%d %B, %Y")
    prev_month_year = prev_month_year[3:]
    prev_month_year = prev_month_year.replace(',', '')
    return prev_month_year

################Function is used to obtain relevant date to insert in EIB#######################
def getPrevMonth():
    first_day_of_this_month = date.today().replace(day=1) 
    last_day_prev_month = first_day_of_this_month - timedelta(days=1) 
    prev_month_name = last_day_prev_month.strftime('%B')
    return prev_month_name 

################Function is used to convert date to a supported format#######################
def applyFormat(ws):
    #Change number to comma-style
    for row in range(1, 100):
        ws["N{}".format(row)].number_format = "MM/DD/YYYY"

#Style results range    
def set_border(ws, cell_range):
    #applying border and alignment
        border = Border(bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell,'border',border)) for cell in flattened]


########################Transfer WMF Fund Expenses Categories to dictionnary in order to use in different function#################
def get_Spend_Categories(ws):
    global spend_category_dic
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4, values_only=True):
        if row[1]:
            #####Condition to only add 'Spend Category Hierarchy Object' titled "WMF FUND EXPENSES"####
            if "WMF" in row[1].upper() and "FUND EXPENSES" in row[1].upper():
                spend_Category_Object, reference_ID = row[0].strip(), row[3].strip()
                spend_category_dic[spend_Category_Object] = reference_ID
    #####Spend Category Data Audit is Organized in the below dictionnary in Key, Values ### 
    #####-----> WMF Fund Exp Audit: WMG_SC_WMF_FUND_EXP_AUDIT
    return spend_category_dic

#########################Gets Column Position of Accruals to Post (Foreign CUrrency and USD)#######################################
def getPos(ws):
    post_Pos, usd_Pos = 0, 0

    #Loops through first three rows to get Accrual Column positions
    for i in range(1,4):
        for col_index,val in enumerate(ws[i]):
            strng = str(val.value)
            if isinstance(strng, str):
                if 'POST' in strng.upper() and 'USD' not in strng.upper():
                    post_Pos = col_index + 1
                elif 'USD' in strng.upper():
                    usd_Pos = col_index + 1
                if post_Pos and usd_Pos:
                    break
    
    ##If Unable to find USD column, add 2 to Post col
    #if not usd_Pos:
     #   usd_Pos = post_Pos + 2
            
    
    a = [post_Pos, usd_Pos]
    print(f"Column positions: {a}")
    return a

def closeMatches(exp_group_name):
    global spend_category_dic
    dic = spend_category_dic
    lst = []

    ##Add dictionnary Keys to a list
    for k,v in dic.items():
        lst.append(k)
    
    #Find string that is in list that is closest to exp_group_name in Accrual file
    possib = get_close_matches(exp_group_name, lst, 1)
    print(f'Close match is {possib}')
    return possib[0]

def processFromAccrual(accrual_ws, eib_ws2):
    global spend_category_dic
    position_list = getPos(accrual_ws)
    row_loc = 4
    post_col, usd_col = position_list[0]-1, position_list[1]-1 #Deduce 1 as tuple indices start at zero
    currency_map = {'USD':0,'CAD':0,'GBP':0,'EUR':0, 'CHF':0,'AED':0,'DKK':0}
    eib_map = {'header_key':2,'line_key':3,'company':5,'ledger':6,'account_set':7,'debit':10,'credit':11,'currency':12\
                        ,'ledger_debit_amount':14,'ledger_credit_amount':15,'memo':16,'cost_center':19,'location':22,'spend_category':25}
    ###############################################Generate Top Half of EIB################################################################
    for row in accrual_ws.iter_rows(min_row=row_loc, max_row=300, min_col=1, max_col=46, values_only=True):
        if row[0]:
            exp_group_name, currency, post_val, usd_val = row[1], row[2], row[post_col], row[usd_col]
            if 'BU_12101' in row[0].upper() and 'TOTAL' not in row[0].upper():
                if row[post_col] != 0 and round(post_val,2) != 0:
                    print(f"{row[post_col]} position: {row_loc}")
                    eib_ws2.cell(row=row_loc, column=eib_map['header_key']).value = 'WMFOFF121'
                    eib_ws2.cell(row=row_loc, column=eib_map['line_key']).value = row_loc - 5
                    eib_ws2.cell(row=row_loc, column=eib_map['company']).value = 'BU_12101'
                    eib_ws2.cell(row=row_loc, column=eib_map['ledger']).value = 62000
                    eib_ws2.cell(row=row_loc, column=eib_map['location']).value = 'LX000'
                    eib_ws2.cell(row=row_loc, column=eib_map['cost_center']).value = 52500
                    eib_ws2.cell(row=row_loc, column=eib_map['currency']).value = currency
                    eib_ws2.cell(row=row_loc, column=eib_map['memo']).value = exp_group_name
                    currency_map['USD'] += round(post_val,2) if "USD" in currency else currency_map['USD']
                    currency_map['CAD'] += round(post_val,2) if "CAD" in currency else currency_map['CAD']
                    currency_map['GBP'] += round(post_val,2) if "GBP" in currency else currency_map['GBP']
                    currency_map['EUR'] += round(post_val,2) if "EUR" in currency else currency_map['EUR']
                    currency_map['CHF'] += round(post_val,2) if "CHF" in currency else currency_map['CHF']
                    currency_map['AED'] += round(post_val,2) if "AED" in currency else currency_map['AED']
                    currency_map['DKK'] += round(post_val,2) if "DKK" in currency else currency_map['DKK']
                    try:
                        eib_ws2.cell(row=row_loc, column=eib_map['spend_category']).value = spend_category_dic[closeMatches(exp_group_name)]
                    except Exception:
                        print(f"Unable to match {exp_group_name} in Accrual to Data Audit, Please Update Data Audit")
                        #continue
                    eib_ws2.cell(row=row_loc, column=eib_map['account_set']).value = 'WMG_FIN_CHILD_ACCOUNT_SET'
                    if round(post_val,2) > 0:
                        eib_ws2.cell(row=row_loc, column=eib_map['debit']).value = round(post_val,2)
                        eib_ws2.cell(row=row_loc, column=eib_map['ledger_debit_amount']).value = round(usd_val,2)
                    elif round(post_val,2) < 0:
                        eib_ws2.cell(row=row_loc, column=eib_map['credit']).value = round(abs(post_val),2)
                        eib_ws2.cell(row=row_loc, column=eib_map['ledger_credit_amount']).value = round(abs(usd_val),2)
                    
                    row_loc += 1
    b = row_loc - 1                
    set_border(eib_ws2, 'A'+ str(b) + ':AT' + str(b))
    print(row_loc)
        ###############################################Generate Bottom Half of EIB################################################################
    bottom_row_loc, i = row_loc, 0
    for cur, value in currency_map.items():
        if value != 0:
            if value > 0:
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['header_key']).value = 'WMFOFF121'
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['line_key']).value = bottom_row_loc - 5
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['company']).value = 'BU_12101'
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['ledger']).value = 21900
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['currency']).value = currency
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['credit']).value = value
                i+=1
            else:
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['header_key']).value = 'WMFOFF121'
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['line_key']).value = bottom_row_loc - 5
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['company']).value = 'BU_12101'
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['ledger']).value = 21900
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['currency']).value = currency
                eib_ws2.cell(row=bottom_row_loc + i, column=eib_map['debit']).value = value
                i+=1

    


def updateImportAccountingTab(ws):
    row_loc = 6
    for row in ws.iter_rows(min_row=row_loc, max_row=6, min_col=1, max_col=23, values_only=True):
        ws.cell(row=row_loc, column=14).value = getPrevMonthEndDate()
        ws.cell(row=row_loc, column=16).value = 'WMF Fixed Admin Fee ' + getPrevMonth()
        row_loc += 1
    applyFormat(ws)

def main(accrualSourcePath, accrualDest):
    system = platform.system() ##Windows or MAC

#####################Part 1: Load and Obtain Spend Categories from DATA Audit file###################
    spend_Category_Path = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/Admin Fee accrual/Data_Audit_-_Spend_Categories.xlsx'\
         if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\000-Common Files\Virtual - GA Lab\cache\Templates\GA LAB Data Audit Spend Category.xlsx'
   
    data_Audit_ws = checkAndLoad_Data_Audit(spend_Category_Path)
    data_Audit_Dict = get_Spend_Categories(data_Audit_ws)

##################Part2: Load WMF Fund Expense Accrual File and EIB Template then write to EIB#####################
    expense_Accrual_Path = accrualSourcePath
    
    expense_Accrual_old = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/Admin Fee accrual/WMF 0220(UPDATED).xlsx'\
         if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\000-Common Files\Virtual - GA Lab\cache\Templates\Data_Audit_-_Spend_Categories.xlsx'

    eib_Source = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/Admin Fee accrual/WLSA EIB Test.xlsx' if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\000-Common Files\Virtual - GA Lab\cache\Templates\Admin Fee WLSA EIB Template.xlsx'
    
    #CHECK EIB FILE IS LOADED PROPERLY
    wbws = checkAndLoad_EIB(eib_Source)

    #CHECK EIB FILE IS LOADED PROPERLY
    expense_Accrual_ws = checkAndLoad_expense(expense_Accrual_Path)

    #EIB Workbook and Worksheets
    eib_wb = wbws[0]
    eib_ws1 = wbws[1]
    eib_ws2 = wbws[2]

    #Generate Import Tab in EIB
    updateImportAccountingTab(eib_ws1)

    #Using Expense Accrual File to Generate EIB Journal Lines 
    processFromAccrual(expense_Accrual_ws, eib_ws2)

    accrualDest = accrualDest +'_EIBUpload.xlsx' if '.xlsx' not in accrualDest else accrualDest.replace('.xlsx', '_EIBUpload.xlsx')
    checkAndSave(eib_wb, accrualDest)

if __name__ == "__main__":
    main(accrualSourcePath, accrualDest)