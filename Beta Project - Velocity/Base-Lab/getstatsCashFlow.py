################################################################################
##
## BY: Sebastien St Vil
##
## This project can be used freely for all uses, any information in the visual
## interface (GUI) can be modified without any implication.
##
##
################################################################################

import platform
import sys as Sys
import pandas as pd
import numpy as np
import matplotlib
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


cellTop, cellBottom = 0, 0
grand_total_pos = 11 #Number chosen so that row corresponds to sheet
Quarter, pTabName = '', ''
strtdate, dataTab = '', 'Memo'
boiteDeCrayon = ['7FD5FF','2F8B1A','F8FA4E','4E64D2','000000','F98107','F90707','FDB4B4', '3CA463'] 
                 #0:bleupale, 1:vert, 2:jaune, 3:bleu, 4:noir, 5:jaune abricot, 6:rouge, 7:rose, 8:lightgreen


def getColorFill(couleur):
    blueFill = PatternFill(start_color=couleur,
                   end_color=couleur,
                   fill_type='solid')
    return blueFill


def writeStatsToFile(nbStat, lbStat, ocfStat, destination):
    try:
        writer = pd.ExcelWriter(destination, engine='openpyxl', mode='a')
        writer.book = load_workbook(destination)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        nbStat.to_excel(writer, sheet_name='NB', freeze_panes=(1,0))
        lbStat.to_excel(writer, sheet_name='LB', freeze_panes=(1,0))
        ocfStat.to_excel(writer, sheet_name='OCF', freeze_panes=(1,0))
        writer.save()
        writer.close()
    except Exception:
        print("Statistic Performance Tabs are not properly saved")

    #############    ####################################
    #checkAndSave(wb, "tgjrigjsdf" + destination)
    ##################When exception is created due to file not existing, file looks appropriate

    #checkAndSave(wb, destination)

def getDataStat(filepath, memoTName):
    
    memo_data_df = pd.read_excel(filepath, sheet_name=memoTName, usecols=['Account Name', 'NB', 'LB', 'OCF'])
    memo_data_df = memo_data_df.dropna()
    df_source = memo_data_df

    df1 = df_source.groupby(['Account Name'])['NB'].sum().sort_values(ascending=False)
    df1.loc['Total'] = df1.sum()

    df2 = df_source.groupby(['Account Name'])['LB'].sum().sort_values(ascending=True)
    df2.loc['Total'] = df2.sum()

    df3 = df_source.groupby(['Account Name'])['OCF'].sum().sort_values(ascending=True)
    df3.loc['Total'] = df3.sum()

    return [df1, df2, df3]

def main(cashflowDest):
    system = platform.system() ##Windows or MAC
    #########################SOURCE FILE LOCATION#################################################
    cashflowDest = cashflowDest +'.xlsx'
    ##Read and Analyze file with Pandas
    stats_pivot_tables = getDataStat(cashflowDest, "Memo")

    nbStat = stats_pivot_tables[0]
    lbStat = stats_pivot_tables[1]
    ocfStat = stats_pivot_tables[2]

    print(nbStat)
    print(lbStat)
    print(ocfStat)

    #checkAndSave(wb, destination)

    #writeStatsToFile(nbStat, lbStat, ocfStat, wb2, destination)


    #Load Updated File and Save
    #checkAndLoad2(wb2, destination)

    #processWorksheet(pivottab)

    #checkAndSave(wb2, destination)

    #writeStatsToFile(nbStat, lbStat, ocfStat, destination)
    return dataTab

if __name__ == "__main__":
    main(cashflowDest)





    
