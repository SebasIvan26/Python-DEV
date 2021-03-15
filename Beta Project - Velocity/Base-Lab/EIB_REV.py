import sys as Sys
import platform
import requests
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from datetime import date, timedelta
import STAT_REV as rev

def getPrevMonthEndDate():
    #ex: 2021-01-31 **Changes to 1/31/21 when inserted into excel
    last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
    #start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
    return last_day_of_prev_month

def getPrevMonthYear():
    #ex: February 2021 --> str
    prev_month_year = date.today().replace(day=1) - timedelta(days=1)
    prev_month_year = prev_month_year.strftime("%d %B, %Y")
    prev_month_year = prev_month_year[3:]
    prev_month_year = prev_month_year.replace(',', '')
    return prev_month_year

def getPrevMonth():
    first_day_of_this_month = date.today().replace(day=1) 
    last_day_prev_month = first_day_of_this_month - timedelta(days=1) 
    prev_month_name = last_day_prev_month.strftime('%B')
    return prev_month_name 

def applyFormat(ws):

    #Change number to comma-style
    for row in range(1, 100):
        ws["N{}".format(row)].number_format = "MM/DD/YYYY"


def getStart(ws):
    dataPlacement = 1 ##Variable to determine where data starts, essentially the placement
    for row in ws.iter_rows(min_row=0, max_row=300, min_col=1, max_col=8, values_only=True):
        if 'BU' in str(row[4]):
            break;
    return dataPlacement + 1

def getWebRevTemplate(source):
    url = "https://github.com/SebasIvan26/testinghello/blob/master/Templates/REV%20EIB%20Template.xlsx?raw=true"

    username = 'kseb0@hotmail.com'
    token =  '538cd0c05c0207b17a43ca6324f12191a73bb696'

    r = requests.get(url,auth=(username,token))
    path = source
    with open(path, 'wb') as f:
        f.write(r.content)
    return path

#Load existing spreadsheet
def checkAndLoad(source):
    #Create active worksheet
    #getWebRevTemplate(source) Downloads Template from github repository -->Disactivated
    try:
        wb = load_workbook(source)
        ws = wb.sheetnames
        
        #Ensuring to load the "WRPP" workbook
        for s in ws:
            if 'Import Account'.upper() in s.upper():
                import_tab = s
            elif 'Journal Entry'.upper() in s.upper():
                entry_tab = s
                #Selecting WRPP Worksheet
        ws = wb[import_tab]
        ws2 = wb[entry_tab]
    #except openpyxl.utils.exceptions.InvalidFileException:
    except Exception:
        print("Unable to load file....Please use Valid file format")
        SystemExit()
    except FileNotFoundError:
        print('Unable to load file....Please enter a valid file path')
        SystemExit()
    print("Source file loaded...")
    
    return [wb, ws, ws2]

def checkAndSave(wb,dest):
    try:
        wb.save(dest)
    except:
        print("Error while saving file")
    
    print("File is successfully saved in specified file path.")

def updateImportAccountingTab(ws):
    row_loc = 6

    for row in ws.iter_rows(min_row=row_loc, max_row=6, min_col=1, max_col=23, values_only=True):
        ws.cell(row=row_loc, column=14).value = getPrevMonthEndDate()
        ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT REVENUE"
        row_loc += 1

# {'GLOBAL/INTERNATIONAL_EQUITY': ['TOK', 1684243478.0225258]}
def updateEntryLines(dic, ws):
    global credit_Col
    global bu_Data_Audit

    row_loc = 6
    for row in ws.iter_rows(min_row=row_loc, max_row=16, min_col=1, max_col=46, values_only=True):
        if row[4]:
            ex_ref_ID = str(row[16]).upper()
            ledger = str(row[5])
            if 'CORE_EQUITY' in ex_ref_ID:
                ws.cell(row=row_loc, column=15).value = round(dic['CORE_EQUITY_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue"
            if 'COMPLEX_ALPHA' in ex_ref_ID:
                ws.cell(row=row_loc, column=15).value = round(dic['COMPLEX_ALPHA_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue"
            if 'SPECIAL_EQUITY' in ex_ref_ID:
                ws.cell(row=row_loc, column=15).value = round(dic['SPECIAL_EQUITY_TOTAL'],2) if 'SPECIALTY_EQUITY_TOTAL' not in dic else\
                    round(dic['SPECIALTY_EQUITY_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue"
            if 'FIXED_INCOME' in ex_ref_ID:
                ws.cell(row=row_loc, column=15).value = round(dic['FIXED_INCOME_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue"
            if 'GLOBAL_INTL' in ex_ref_ID:
                ws.cell(row=row_loc, column=15).value = round(dic['GLOBAL/INTERNATIONAL_EQUITY_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue"
            if '79993' in ledger:
                ws.cell(row=row_loc, column=14).value = round(dic['GRAND_TOTAL'],2)
                ws.cell(row=row_loc, column=16).value = getPrevMonth() + " STAT Revenue - Contra Revenue"
        row_loc += 1

def main(dic, bucketDestPath):
    system = platform.system() ##Windows or MAC
    #########################SOURCE FILE LOCATION#################################################

    source = r'/Users/sebastienstvil/Documents/Python/Python-DEV/Beta Project - Velocity/Base-Lab/Testing/bucket_Tests/Original EIB/Templates/REV EIB Template.xlsx' if system == 'Darwin' else r'\\prod-corpfile\netshare\GA\000-Common Files\Virtual - GA Lab\cache\Templates\REV EIB Template.xlsx'

    #########################DESTINATION FILE LOCATION#################################################
    destination = source

    #CHECK FILE IS LOADED PROPERLY
    wbws = checkAndLoad(source)

    wb = wbws[0]
    ws1 = wbws[1]
    ws2 = wbws[2]


    #INSERT EXCEL NUMBERS INTO DATA STRUCTURES TO ORGANIZE (ORGANIZED BY POOLS)
    updateImportAccountingTab(ws1)
    updateEntryLines(dic, ws2)
    applyFormat(ws1)
    #dicto = processFromSheet1(ws1, dic)
    bucketDestPath = bucketDestPath +'_REV_EIBUpload.xlsx' if '.xlsx' not in bucketDestPath else bucketDestPath.replace('.xlsx', '_REV_EIBUpload.xlsx')
    checkAndSave(wb, bucketDestPath)


if __name__ == "__main__":
    main(dic, bucketDestPath)
    